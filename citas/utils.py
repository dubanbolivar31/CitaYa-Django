import io
import os
import resend
from datetime import date

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

from django.template.loader import render_to_string
from django.conf import settings


# =========================================================================
# CORREO — Resend
# =========================================================================

def enviar_confirmacion_cita(paciente, cita):
    context      = {'paciente': paciente, 'cita': cita}
    html_content = render_to_string('emails/confirmacion_cita.html', context)
    try:
        resend.Emails.send({
            "from":    settings.DEFAULT_FROM_EMAIL,
            "to":      [paciente.correo],
            "subject": f"Confirmación de Cita - {cita.cita}",
            "html":    html_content,
        })
    except Exception as e:
        print(f"Error al enviar el correo: {e}")


# =========================================================================
# PALETA
# =========================================================================

C_ACCENT   = colors.HexColor('#1a56db')
C_GREEN    = colors.HexColor('#059669')
C_TEXT     = colors.HexColor('#0f172a')
C_TEXT2    = colors.HexColor('#475569')
C_TEXT3    = colors.HexColor('#94a3b8')
C_BORDER   = colors.HexColor('#e2eaf3')
C_ROW_ODD  = colors.HexColor('#f4f7fb')
C_ROW_EVEN = colors.white
C_WHITE    = colors.white

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'img', 'logo.png')


# =========================================================================
# TEMPLATE — header limpio + footer sutil
# =========================================================================

class _ReportDoc(BaseDocTemplate):

    def __init__(self, buffer, page_size, report_title):
        self.report_title = report_title
        SIDE = 1.8 * cm
        TOP  = 2.8 * cm
        BOT  = 1.6 * cm

        super().__init__(
            buffer,
            pagesize=page_size,
            leftMargin=SIDE,
            rightMargin=SIDE,
            topMargin=TOP,
            bottomMargin=BOT,
        )
        frame = Frame(SIDE, BOT, page_size[0] - 2 * SIDE, page_size[1] - TOP - BOT, id='main')
        self.addPageTemplates([
            PageTemplate(id='citaya', frames=[frame], onPage=self._draw_page)
        ])

    def _draw_page(self, canv, doc):
        canv.saveState()
        W    = doc.pagesize[0]
        H    = doc.pagesize[1]
        SIDE = 1.8 * cm

        header_y = H - 1.5 * cm
        logo_h   = 0.68 * cm

        logo_drawn_w = 0
        if os.path.exists(LOGO_PATH):
            try:
                from reportlab.lib.utils import ImageReader
                img = ImageReader(LOGO_PATH)
                iw, ih = img.getSize()
                logo_w = (iw / ih) * logo_h
                canv.drawImage(
                    LOGO_PATH, SIDE, header_y - logo_h * 0.15,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto',
                )
                logo_drawn_w = logo_w
            except Exception:
                logo_drawn_w = 0

        canv.setFont('Helvetica-Bold', 10.5)
        canv.setFillColor(C_TEXT)
        canv.drawCentredString(W / 2, header_y, self.report_title)

        canv.setFont('Helvetica', 8)
        canv.setFillColor(C_TEXT3)
        canv.drawRightString(W - SIDE, header_y, date.today().strftime('%d / %m / %Y'))

        sep_y = header_y - 0.38 * cm
        canv.setStrokeColor(C_ACCENT)
        canv.setLineWidth(0.7)
        canv.line(SIDE, sep_y, W - SIDE, sep_y)

        foot_y = 0.72 * cm
        canv.setStrokeColor(C_BORDER)
        canv.setLineWidth(0.4)
        canv.line(SIDE, foot_y, W - SIDE, foot_y)

        canv.setFont('Helvetica', 7)
        canv.setFillColor(C_TEXT3)
        canv.drawString(SIDE, foot_y - 0.3 * cm, 'Documento de uso interno · CitaYa')
        canv.drawRightString(W - SIDE, foot_y - 0.3 * cm, f'Página {doc.page}')

        canv.restoreState()


# =========================================================================
# BUILDER DE TABLA PDF
# =========================================================================

def _build_pdf_table(df: pd.DataFrame, title: str, page_size=A4) -> bytes:
    buffer = io.BytesIO()
    doc = _ReportDoc(buffer, page_size=page_size, report_title=title)

    styles = getSampleStyleSheet()

    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8,
        textColor=C_TEXT, leading=11,
    )
    hdr_style = ParagraphStyle(
        'Hdr', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=C_WHITE, leading=11,
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8,
        textColor=C_TEXT3, spaceAfter=8,
    )

    elements = []
    elements.append(Spacer(1, 0.15 * cm))
    elements.append(Paragraph(f'{len(df)} registro(s)', meta_style))

    headers   = list(df.columns)
    hdr_row   = [Paragraph(h, hdr_style) for h in headers]
    data_rows = [hdr_row]

    for _, row in df.iterrows():
        data_rows.append([
            Paragraph(str(v) if v is not None else '—', cell_style)
            for v in row.values
        ])

    page_w = page_size[0] - 3.6 * cm
    col_w  = page_w / len(headers)

    table = Table(data_rows, colWidths=[col_w] * len(headers), repeatRows=1)

    cmds = [
        ('BACKGROUND',    (0, 0), (-1, 0),  C_ACCENT),
        ('TOPPADDING',    (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  9),
        ('LEFTPADDING',   (0, 0), (-1, 0),  8),
        ('RIGHTPADDING',  (0, 0), (-1, 0),  8),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('VALIGN',        (0, 0), (-1, 0),  'MIDDLE'),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.5, C_GREEN),
        ('TOPPADDING',    (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 7),
        ('LEFTPADDING',   (0, 1), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 1), (-1, -1), 8),
        ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (0, 1), (-1, -1), 'LEFT'),
        ('LINEBELOW',     (0, 1), (-1, -1), 0.35, C_BORDER),
        ('BOX',           (0, 0), (-1, -1), 0.5,  C_BORDER),
    ]

    for i in range(1, len(data_rows)):
        bg = C_ROW_ODD if i % 2 == 1 else C_ROW_EVEN
        cmds.append(('BACKGROUND', (0, i), (-1, i), bg))

    table.setStyle(TableStyle(cmds))
    elements.append(table)

    doc.multiBuild(elements)
    buffer.seek(0)
    return buffer.read()


def _queryset_to_df(queryset, fields, rename=None):
    records = list(queryset.values(*fields))
    df = pd.DataFrame(records) if records else pd.DataFrame(columns=fields)
    if rename:
        df = df.rename(columns=rename)
    return df


# =========================================================================
# GENERADORES PÚBLICOS
# =========================================================================

def generar_df_administradores(queryset):
    fields = ['id_admin', 'tipo_doc', 'numero_doc', 'nombre', 'apellido',
              'genero', 'telefono', 'correo', 'estado']
    rename = {
        'id_admin': 'ID', 'tipo_doc': 'Tipo Doc', 'numero_doc': 'Núm. Doc',
        'nombre': 'Nombre', 'apellido': 'Apellido', 'genero': 'Género',
        'telefono': 'Teléfono', 'correo': 'Correo', 'estado': 'Estado',
    }
    df = _queryset_to_df(queryset, fields, rename)
    if 'Estado' in df.columns:
        df['Estado'] = df['Estado'].map({True: 'Activo', False: 'Inactivo'})
    return df

def generar_pdf_administradores(queryset):
    return _build_pdf_table(generar_df_administradores(queryset), 'Reporte de Administradores')

def generar_excel_administradores(queryset):
    return _df_to_excel(generar_df_administradores(queryset), 'Administradores')


def generar_df_medicos(queryset):
    fields = ['id_medico', 'tipo_doc', 'numero_doc', 'nombre', 'apellido',
              'genero', 'especialidad', 'telefono', 'correo', 'estado']
    rename = {
        'id_medico': 'ID', 'tipo_doc': 'Tipo Doc', 'numero_doc': 'Núm. Doc',
        'nombre': 'Nombre', 'apellido': 'Apellido', 'genero': 'Género',
        'especialidad': 'Especialidad', 'telefono': 'Teléfono',
        'correo': 'Correo', 'estado': 'Estado',
    }
    df = _queryset_to_df(queryset, fields, rename)
    if 'Estado' in df.columns:
        df['Estado'] = df['Estado'].map({True: 'Activo', False: 'Inactivo'})
    return df

def generar_pdf_medicos(queryset):
    return _build_pdf_table(generar_df_medicos(queryset), 'Reporte de Médicos', page_size=landscape(A4))

def generar_excel_medicos(queryset):
    return _df_to_excel(generar_df_medicos(queryset), 'Médicos')


def generar_df_pacientes(queryset):
    fields = ['id_paciente', 'tipo_doc', 'numero_doc', 'nombre', 'apellido',
              'genero', 'fecha_nacimiento', 'tipo_sangre', 'direccion',
              'telefono', 'correo', 'estado']
    rename = {
        'id_paciente': 'ID', 'tipo_doc': 'Tipo Doc', 'numero_doc': 'Núm. Doc',
        'nombre': 'Nombre', 'apellido': 'Apellido', 'genero': 'Género',
        'fecha_nacimiento': 'Nacimiento', 'tipo_sangre': 'Sangre',
        'direccion': 'Dirección', 'telefono': 'Teléfono',
        'correo': 'Correo', 'estado': 'Estado',
    }
    df = _queryset_to_df(queryset, fields, rename)
    if 'Estado' in df.columns:
        df['Estado'] = df['Estado'].map({True: 'Activo', False: 'Inactivo'})
    return df

def generar_pdf_pacientes(queryset):
    return _build_pdf_table(generar_df_pacientes(queryset), 'Reporte de Pacientes', page_size=landscape(A4))

def generar_excel_pacientes(queryset):
    return _df_to_excel(generar_df_pacientes(queryset), 'Pacientes')


def generar_df_agendamientos(queryset):
    from django.db.models import F, Value
    from django.db.models.functions import Concat
    qs = queryset.annotate(
        nombre_paciente=Concat(F('id_paciente__nombre'), Value(' '), F('id_paciente__apellido')),
        nombre_medico=Concat(F('id_medico__nombre'), Value(' '), F('id_medico__apellido')),
    )
    fields = ['id_agendamiento', 'cita', 'fecha', 'hora', 'nombre_paciente', 'nombre_medico']
    rename = {
        'id_agendamiento': 'ID', 'cita': 'Tipo de Cita', 'fecha': 'Fecha',
        'hora': 'Hora', 'nombre_paciente': 'Paciente', 'nombre_medico': 'Médico',
    }
    df = _queryset_to_df(qs, fields, rename)
    if 'Hora' in df.columns:
        df['Hora'] = df['Hora'].apply(
            lambda h: h.strftime('%H:%M') if hasattr(h, 'strftime') else str(h)
        )
    return df

def generar_pdf_agendamientos(queryset):
    return _build_pdf_table(generar_df_agendamientos(queryset), 'Reporte de Agendamientos')

def generar_excel_agendamientos(queryset):
    return _df_to_excel(generar_df_agendamientos(queryset), 'Agendamientos')


def generar_df_historial(queryset):
    from django.db.models import F, Value
    from django.db.models.functions import Concat
    qs = queryset.annotate(
        nombre_paciente=Concat(F('id_paciente__nombre'), Value(' '), F('id_paciente__apellido')),
        nombre_medico=Concat(F('id_medico__nombre'), Value(' '), F('id_medico__apellido')),
    )
    fields = ['id_historial', 'fecha_creacion', 'nombre_paciente', 'nombre_medico', 'antecedentes']
    rename = {
        'id_historial': 'ID', 'fecha_creacion': 'Fecha',
        'nombre_paciente': 'Paciente', 'nombre_medico': 'Médico Tratante',
        'antecedentes': 'Antecedentes',
    }
    return _queryset_to_df(qs, fields, rename)

def generar_pdf_historial(queryset):
    return _build_pdf_table(generar_df_historial(queryset), 'Reporte de Historial Clínico')

def generar_excel_historial(queryset):
    return _df_to_excel(generar_df_historial(queryset), 'Historial Clínico')


# =========================================================================
# EXCEL
# =========================================================================

def _df_to_excel(df: pd.DataFrame, sheet_name: str = 'Reporte') -> bytes:
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        num_cols = len(df.columns)
        last_col = get_column_letter(num_cols)

        fill_accent = PatternFill('solid', fgColor='1A56DB')
        fill_odd    = PatternFill('solid', fgColor='F4F7FB')
        fill_even   = PatternFill('solid', fgColor='FFFFFF')

        thin  = Side(style='thin',   color='E2EAF3')
        green = Side(style='medium', color='059669')

        border_data   = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_header = Border(left=thin, right=thin, top=thin, bottom=green)

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        right  = Alignment(horizontal='right',  vertical='center')

        ws.merge_cells(f'A1:{get_column_letter(max(num_cols - 1, 1))}1')
        t = ws['A1']
        t.value     = f'Reporte de {sheet_name}'
        t.font      = Font(bold=True, size=12, color='0F172A', name='Calibri')
        t.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 26

        d = ws.cell(row=1, column=num_cols)
        d.value     = date.today().strftime('%d/%m/%Y')
        d.font      = Font(size=8, color='94A3B8', name='Calibri')
        d.alignment = right

        ws.row_dimensions[2].height = 3
        for col in range(1, num_cols + 1):
            ws.cell(row=2, column=col).fill = PatternFill('solid', fgColor='E2EAF3')

        for cell in ws[3]:
            cell.fill      = fill_accent
            cell.font      = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
            cell.alignment = center
            cell.border    = border_header
        ws.row_dimensions[3].height = 22

        for row_idx in range(4, 4 + len(df)):
            fill = fill_odd if row_idx % 2 == 0 else fill_even
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill      = fill
                cell.alignment = left
                cell.border    = border_data
                cell.font      = Font(size=9, color='0F172A', name='Calibri')
            ws.row_dimensions[row_idx].height = 18

        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = 'A4'
        ws.sheet_properties.tabColor = '1A56DB'

    buffer.seek(0)
    return buffer.read()